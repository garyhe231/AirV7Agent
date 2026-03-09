import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from datetime import datetime
import shutil
import os

TEMPLATE_PATH = "/Users/jhe/Downloads/[Client]_ Air Template_v7.xlsx"


def export_to_v7(bid_data: Dict[str, Any], output_path: str) -> str:
    """Copy the V7 template and populate it with bid data."""
    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path)

    _fill_internal_pricing(wb, bid_data)
    _fill_kickoff(wb, bid_data)

    wb.save(output_path)
    return output_path


def _fill_internal_pricing(wb, bid_data: Dict[str, Any]):
    ws = wb["Internal Pricing"]
    lanes = bid_data.get("lanes", [])

    # Data starts at row 7 (rows 1-6 are headers/legend)
    for i, lane in enumerate(lanes):
        row = 7 + i
        col_map = {
            1: i + 1,                                          # Lane ID
            2: lane.get("pricing_request_id", ""),            # Pricing Request ID
            9: lane.get("client_name", bid_data.get("client_name", "")),  # Client
            10: lane.get("effective_date", ""),               # Effective Date
            11: lane.get("expiration_date", ""),              # Expiration Date
            12: lane.get("incoterms", ""),                    # Incoterms
            18: lane.get("origin_country", ""),               # Origin Country
            20: lane.get("origin_city", ""),                  # Origin City
            23: lane.get("origin_airport", ""),               # Origin Airport
            25: lane.get("destination_country", ""),          # Dest Country
            27: lane.get("destination_city", ""),             # Dest City
            30: lane.get("destination_airport", ""),          # Dest Airport
            32: lane.get("total_shipments", ""),              # Total Shipments
            36: lane.get("chargeable_weight_kg", ""),         # Chargeable Weight
            38: lane.get("avg_shipment_kg", ""),              # Avg Shipment
            49: lane.get("service_tier", ""),                 # Service Tier
            51: lane.get("origin_currency", "USD"),           # Origin Currency
            52: lane.get("pickup_min", ""),                   # Pickup Min
            55: lane.get("pickup_kg_1000", ""),               # Pickup 1000kg
            60: lane.get("origin_thc", ""),                   # Origin THC
            61: lane.get("screening", ""),                    # Screening
            62: lane.get("doc_fee", ""),                      # Doc Fee
            63: lane.get("export_customs", ""),               # Export Customs
            65: lane.get("main_currency", "USD"),             # Main Currency
            73: lane.get("buy_rate_per_kg", ""),              # Buy 1000kg
            75: lane.get("fuel_surcharge", ""),               # FSC
            76: lane.get("security_charge", ""),              # SSC
            77: lane.get("ams_ens", ""),                      # AMS/ENS
            79: lane.get("pss", ""),                          # PSS
            82: lane.get("airline", ""),                      # Airline
            83: lane.get("routing", ""),                      # Routing
            86: lane.get("transit_min", ""),                  # Transit Min
            87: lane.get("transit_max", ""),                  # Transit Max
            88: lane.get("dest_currency", "USD"),             # Dest Currency
            89: lane.get("delivery_min", ""),                 # Delivery Min
            92: lane.get("delivery_kg_1000", ""),             # Delivery 1000kg
            97: lane.get("dest_thc", ""),                     # Dest THC
            98: lane.get("import_service", ""),               # Import Service
            99: lane.get("import_customs", ""),               # Import Customs
            100: lane.get("doc_turnover", ""),                # Doc Turnover
            118: lane.get("sell_rate_per_kg", ""),            # Sell 1000kg
        }
        for col, value in col_map.items():
            if value != "":
                ws.cell(row=row, column=col).value = value


def _fill_kickoff(wb, bid_data: Dict[str, Any]):
    if "kickoff" not in bid_data or not bid_data["kickoff"]:
        return
    ws = wb["Kick Off Questions"]
    kickoff = bid_data["kickoff"]

    # Answer column is B (col 2); questions start at row 3
    answer_map = {
        3: kickoff.get("client_due_date"),
        4: kickoff.get("internal_due_date"),
        5: kickoff.get("commercial_context"),
        6: kickoff.get("opportunity_strategy"),
        7: kickoff.get("expected_rounds"),
        8: kickoff.get("num_providers"),
        9: kickoff.get("client_transit_times"),
        11: kickoff.get("financial_penalties"),
        12: kickoff.get("fuel_policy"),
        13: kickoff.get("cargo_density"),
        14: kickoff.get("accepts_pss"),
        15: kickoff.get("routing_requirements"),
        16: kickoff.get("special_delivery"),
        17: kickoff.get("pricing_instructions"),
        19: kickoff.get("award_methodology"),
        20: kickoff.get("pricing_assessment_logic"),
        21: kickoff.get("what_makes_us_win"),
        22: kickoff.get("other_notes"),
    }
    for row, value in answer_map.items():
        if value:
            ws.cell(row=row, column=2).value = value


def _fmt(val, suffix="") -> str:
    """Format a numeric value or return dash."""
    if val is None or val == "" or val == 0:
        return "-"
    try:
        return f"{float(val):.3f}{suffix}"
    except (TypeError, ValueError):
        return str(val)


def generate_customer_quote_html(bid_data: Dict[str, Any]) -> str:
    """Generate a customer-facing HTML quote with all weight breaks per V7 format."""
    client_name = bid_data.get("client_name", "Client")
    pr_id = bid_data.get("pricing_request_id", "")
    lanes = bid_data.get("lanes", [])

    # Shipment info rows
    shipment_rows = ""
    rate_rows = ""

    for lane in lanes:
        origin = f"{lane.get('origin_city', '')} ({lane.get('origin_airport', '')})"
        dest = f"{lane.get('destination_city', '')} ({lane.get('destination_airport', '')})"
        tier = lane.get("service_tier", "-")
        dg = lane.get("dangerous_goods", "General Cargo") or "General Cargo"
        stackable = lane.get("stackable", "-") or "-"
        packaging = lane.get("packaging", "-") or "-"
        incoterms = lane.get("incoterms", "-") or "-"
        transit = f"{lane.get('transit_min', '?')}–{lane.get('transit_max', '?')} days"
        effective = lane.get("effective_date", "-") or "-"
        expiration = lane.get("expiration_date", "-") or "-"
        port_pair = f"{lane.get('origin_airport', '?')} – {lane.get('destination_airport', '?')}"

        shipment_rows += f"""
        <tr>
            <td>{lane.get('lane_id')}</td>
            <td>{port_pair}</td>
            <td>{origin}</td>
            <td>{dest}</td>
            <td><strong>{tier}</strong></td>
            <td>{dg}</td>
            <td>{stackable}</td>
            <td>{packaging}</td>
            <td>{incoterms}</td>
            <td>{transit}</td>
            <td>{effective}</td>
            <td>{expiration}</td>
        </tr>"""

        # All 7 weight breaks
        s_min = _fmt(lane.get("sell_rate_min"))
        s_45  = _fmt(lane.get("sell_rate_45"))
        s_100 = _fmt(lane.get("sell_rate_100"))
        s_300 = _fmt(lane.get("sell_rate_300"))
        s_500 = _fmt(lane.get("sell_rate_500"))
        s_1k  = _fmt(lane.get("sell_rate_per_kg"))
        s_2k  = _fmt(lane.get("sell_rate_2000"))
        fsc   = _fmt(lane.get("fuel_surcharge"))
        ssc   = _fmt(lane.get("security_charge"))
        ams   = _fmt(lane.get("ams_ens"))
        pss   = _fmt(lane.get("pss"))
        acas  = _fmt(lane.get("acas"))
        ccy   = lane.get("main_currency") or "USD"

        rate_rows += f"""
        <tr>
            <td>{lane.get('lane_id')}</td>
            <td>{port_pair}</td>
            <td><strong>{tier}</strong></td>
            <td>{ccy}</td>
            <td>{s_min}</td>
            <td>{s_45}</td>
            <td>{s_100}</td>
            <td>{s_300}</td>
            <td>{s_500}</td>
            <td><strong style="color:#1a73e8">{s_1k}</strong></td>
            <td>{s_2k}</td>
            <td>{fsc}</td>
            <td>{ssc}</td>
            <td>{ams}</td>
            <td>{pss}</td>
            <td>{acas}</td>
        </tr>"""

    style = """
    <style>
      .quote-export { font-family: Arial, sans-serif; font-size: 12px; color: #1a1a2e; max-width: 100%; }
      .quote-export h2 { font-size: 18px; color: #1a1a2e; margin-bottom: 4px; }
      .quote-meta { color: #666; font-size: 11px; margin-bottom: 16px; }
      .quote-section-title { font-size: 13px; font-weight: 700; color: #1a73e8; margin: 18px 0 8px; border-bottom: 2px solid #1a73e8; padding-bottom: 4px; }
      .quote-table { width: 100%; border-collapse: collapse; margin-bottom: 12px; }
      .quote-table th { background: #1a1a2e; color: #fff; padding: 7px 10px; font-size: 11px; text-align: left; }
      .quote-table td { padding: 6px 10px; border-bottom: 1px solid #e0e0e0; }
      .quote-table tr:nth-child(even) td { background: #f8f9ff; }
      .terms { font-size: 11px; color: #777; margin-top: 14px; line-height: 1.8; border-top: 1px solid #ddd; padding-top: 10px; }
    </style>"""

    return f"""
    {style}
    <div class="quote-export">
        <h2>Airfreight Quote — {client_name}</h2>
        <p class="quote-meta">Pricing Request: {pr_id} &nbsp;|&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d')} &nbsp;|&nbsp; Flexport Air Freight</p>

        <div class="quote-section-title">Shipment Information</div>
        <table class="quote-table">
            <thead>
                <tr>
                    <th>#</th><th>Port Pair</th><th>Origin</th><th>Destination</th><th>Service</th>
                    <th>DG Status</th><th>Stackable</th><th>Packaging</th><th>Incoterms</th>
                    <th>Transit</th><th>Valid From</th><th>Valid To</th>
                </tr>
            </thead>
            <tbody>{shipment_rows}</tbody>
        </table>

        <div class="quote-section-title">Airfreight Rates (per kg, by weight break)</div>
        <table class="quote-table">
            <thead>
                <tr>
                    <th>#</th><th>Port Pair</th><th>Service</th><th>CCY</th>
                    <th>Min</th><th>+45kg</th><th>+100kg</th><th>+300kg</th><th>+500kg</th>
                    <th>+1000kg</th><th>+2000kg</th>
                    <th>FSC/kg</th><th>SSC/kg</th><th>AMS/ENS</th><th>PSS/kg</th><th>ACAS</th>
                </tr>
            </thead>
            <tbody>{rate_rows}</tbody>
        </table>

        <div class="terms">
            <p>* All rates are in the currency shown. Flexport reserves the right to apply a Currency Adjustment Factor (CAF).</p>
            <p>* Offer is valid for acceptance maximum 7 days after submission date.</p>
            <p>* Chargeable weight = MAX(actual weight, volumetric weight). Volumetric = CBM × 166.67 (1:6 ratio).</p>
            <p>* PSS (Peak Season Surcharge) applicable Sep–Dec unless otherwise stated.</p>
            <p>* AMS applicable on US-inbound lanes only. ENS applicable on EU-inbound lanes only.</p>
        </div>
    </div>
    """
